from io import BytesIO
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from fastapi import UploadFile, HTTPException


def create_excel_template() -> BytesIO:
    """Создает Excel шаблон для импорта задач"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Tasks Template"

    # Заголовки колонок
    headers = [
        "Название*",
        "Описание",
        "Ссылка на Идеичную",
        "Исполнитель",
        "Приоритет",
        "Value (1-5)",
        "Reach (1-5)",
        "Budget Impact (1, 1.2, 1.5, 2)",
        "Confidence (0-100)",
        "Важно (0/1)",
        "Срочно (0/1)",
        "Проект (название)"
    ]

    # Стилизация заголовков
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Установка ширины колонок
    column_widths = [25, 40, 25, 20, 15, 15, 15, 25, 20, 15, 15, 20]
    for col_num, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(col_num)].width = width

    # Добавление примера заполнения (закомментированная строка)
    example_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    example_row = 2
    example_data = [
        "Пример задачи",
        "Описание задачи",
        "https://example.com/task",
        "Иван Иванов",
        "High",
        "4",
        "5",
        "1.5",
        "80",
        "1",
        "0",
        "Copilot"
    ]

    for col_num, value in enumerate(example_data, start=1):
        cell = sheet.cell(row=example_row, column=col_num)
        cell.value = value
        cell.fill = example_fill
        cell.alignment = Alignment(wrap_text=True)

    # Добавление инструкций на второй лист
    instructions_sheet = workbook.create_sheet("Инструкция")
    instructions = [
        ["Инструкция по заполнению шаблона задач"],
        [],
        ["Обязательные поля:"],
        ["  • Название* - обязательное поле"],
        [],
        ["Необязательные поля:"],
        ["  • Описание - текстовое описание задачи"],
        ["  • Ссылка на Идеичную - URL задачи"],
        ["  • Исполнитель - имя ответственного"],
        ["  • Приоритет - Low, Medium, High или Highest"],
        [],
        ["Priority Score (Rice Scoring):"],
        ["  • Value - влияние фичи (1-5)"],
        ["  • Reach - охват пользователей (1-5)"],
        ["  • Budget Impact - влияние на бюджет (1, 1.2, 1.5, 2)"],
        ["  • Confidence - уверенность в оценке (0-100)"],
        [],
        ["Матрица Эйзенхауэра:"],
        ["  • Важно - 0 (нет) или 1 (да)"],
        ["  • Срочно - 0 (нет) или 1 (да)"],
        [],
        ["Проект:"],
        ["  • Укажите название существующего проекта"],
        ["  • Если проекта не существует, создайте его в приложении"],
        [],
        ["Примечание:"],
        ["  • Удалите строку с примером перед импортом"],
        ["  • Все поля кроме 'Название' необязательны"],
        ["  • При ошибке импорт остановится с указанием проблемы"]
    ]

    for row_num, row_data in enumerate(instructions, start=1):
        for col_num, value in enumerate(row_data, start=1):
            cell = instructions_sheet.cell(row=row_num, column=col_num)
            cell.value = value
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
            elif "•" in str(value):
                cell.font = Font(size=10)

    instructions_sheet.column_dimensions['A'].width = 80

    # Сохранение в BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


async def parse_excel_file(file: UploadFile) -> List[Dict[str, Any]]:
    """Парсит Excel файл и возвращает список задач для создания"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Неверный формат файла. Разрешены только .xlsx и .xls файлы"
        )

    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents))
        sheet = workbook.active

        # Проверка заголовков
        expected_headers = [
            "Название*", "Описание", "Ссылка на Идеичную", "Исполнитель",
            "Приоритет", "Value (1-5)", "Reach (1-5)", "Budget Impact (1, 1.2, 1.5, 2)",
            "Confidence (0-100)", "Важно (0/1)", "Срочно (0/1)", "Проект (название)"
        ]

        actual_headers = [cell.value for cell in sheet[1]]
        if actual_headers != expected_headers:
            raise HTTPException(
                status_code=400,
                detail="Неверная структура файла. Пожалуйста, используйте шаблон."
            )

        tasks = []
        errors = []

        # Парсинг строк (начиная со 2-й, пропуская заголовок)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Пропускаем пустые строки
            if not any(row):
                continue

            # Пропускаем строку с примером
            if row[0] == "Пример задачи":
                continue

            try:
                # Обязательное поле - название
                if not row[0]:
                    errors.append(f"Строка {row_num}: Название задачи обязательно")
                    continue

                task_data = {
                    "title": str(row[0]).strip(),
                    "description": str(row[1]).strip() if row[1] else None,
                    "link": str(row[2]).strip() if row[2] else None,
                    "assignee": str(row[3]).strip() if row[3] else None,
                    "priority": None,
                    "value": None,
                    "reach": None,
                    "budget_impact": None,
                    "confidence": None,
                    "is_important": None,
                    "is_urgent": None,
                    "project_name": None
                }

                # Приоритет
                if row[4]:
                    priority = str(row[4]).strip()
                    if priority not in ["Low", "Medium", "High", "Highest"]:
                        errors.append(f"Строка {row_num}: Неверный приоритет '{priority}'. Допустимы: Low, Medium, High, Highest")
                        continue
                    task_data["priority"] = priority

                # Value (1-5)
                if row[5] is not None:
                    try:
                        value = int(row[5])
                        if value < 1 or value > 5:
                            errors.append(f"Строка {row_num}: Value должно быть от 1 до 5")
                            continue
                        task_data["value"] = value
                    except (ValueError, TypeError):
                        errors.append(f"Строка {row_num}: Value должно быть числом от 1 до 5")
                        continue

                # Reach (1-5)
                if row[6] is not None:
                    try:
                        reach = int(row[6])
                        if reach < 1 or reach > 5:
                            errors.append(f"Строка {row_num}: Reach должно быть от 1 до 5")
                            continue
                        task_data["reach"] = reach
                    except (ValueError, TypeError):
                        errors.append(f"Строка {row_num}: Reach должно быть числом от 1 до 5")
                        continue

                # Budget Impact (1, 1.2, 1.5, 2)
                if row[7] is not None:
                    try:
                        budget = float(row[7])
                        if budget not in [1.0, 1.2, 1.5, 2.0]:
                            errors.append(f"Строка {row_num}: Budget Impact должно быть 1, 1.2, 1.5 или 2")
                            continue
                        task_data["budget_impact"] = budget
                    except (ValueError, TypeError):
                        errors.append(f"Строка {row_num}: Budget Impact должно быть числом (1, 1.2, 1.5 или 2)")
                        continue

                # Confidence (0-100)
                if row[8] is not None:
                    try:
                        confidence = int(row[8])
                        if confidence < 0 or confidence > 100:
                            errors.append(f"Строка {row_num}: Confidence должно быть от 0 до 100")
                            continue
                        task_data["confidence"] = confidence
                    except (ValueError, TypeError):
                        errors.append(f"Строка {row_num}: Confidence должно быть числом от 0 до 100")
                        continue

                # Важно (0/1)
                if row[9] is not None:
                    try:
                        is_important = int(row[9])
                        if is_important not in [0, 1]:
                            errors.append(f"Строка {row_num}: Важно должно быть 0 или 1")
                            continue
                        task_data["is_important"] = is_important
                    except (ValueError, TypeError):
                        errors.append(f"Строка {row_num}: Важно должно быть 0 или 1")
                        continue

                # Срочно (0/1)
                if row[10] is not None:
                    try:
                        is_urgent = int(row[10])
                        if is_urgent not in [0, 1]:
                            errors.append(f"Строка {row_num}: Срочно должно быть 0 или 1")
                            continue
                        task_data["is_urgent"] = is_urgent
                    except (ValueError, TypeError):
                        errors.append(f"Строка {row_num}: Срочно должно быть 0 или 1")
                        continue

                # Проект
                if row[11]:
                    task_data["project_name"] = str(row[11]).strip()

                tasks.append(task_data)

            except Exception as e:
                errors.append(f"Строка {row_num}: Неожиданная ошибка - {str(e)}")
                continue

        # Если есть ошибки, останавливаем импорт
        if errors:
            error_message = "Ошибки при импорте:\n" + "\n".join(errors)
            raise HTTPException(status_code=400, detail=error_message)

        if not tasks:
            raise HTTPException(
                status_code=400,
                detail="В файле не найдено задач для импорта"
            )

        return tasks

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Ошибка при чтении файла: {str(e)}"
        )

